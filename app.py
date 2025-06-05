from flask import Flask, render_template, request, redirect, url_for, flash
import os
import re
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = Flask(__name__)
app.secret_key = "resolve_secreto"

EXCEL_PATH = 'clientes.xlsx'
PLANILHA_CAMPOS = [
    "Nome", "Notes", "E-mail", "Celular", "Endereço", "Cidade", "CEP",
    "Observações", "Contato", "Tipo de Pessoa", "Data/Hora Cadastro"
]

def formatar_nome(nome):
    return ' '.join([p.capitalize() for p in nome.strip().split()])

def formatar_complemento(texto):
    return texto.strip().capitalize()

def formatar_observacao(obs):
    return obs.strip().capitalize()

def formatar_celular(cel):
    digitos = ''.join(filter(str.isdigit, cel))
    if len(digitos) == 11:
        return f"({digitos[:2]}) {digitos[2:7]}-{digitos[7:]}"
    elif len(digitos) == 10:
        return f"({digitos[:2]}) {digitos[2:6]}-{digitos[6:]}"
    else:
        return cel

def celular_para_wa_link(cel):
    digitos = ''.join(filter(str.isdigit, cel))
    if digitos.startswith("0"):
        digitos = digitos[1:]
    if len(digitos) >= 10:
        return f"https://wa.me/55{digitos}"
    return ""

def validar_cpf(cpf):
    cpf = re.sub(r'\D', '', cpf)
    if len(cpf) != 11 or cpf == cpf[0]*11:
        return False
    soma1 = sum(int(cpf[i]) * (10 - i) for i in range(9))
    dig1 = ((soma1 * 10) % 11) % 10
    soma2 = sum(int(cpf[i]) * (11 - i) for i in range(10))
    dig2 = ((soma2 * 10) % 11) % 10
    return dig1 == int(cpf[9]) and dig2 == int(cpf[10])

def validar_cnpj(cnpj):
    cnpj = re.sub(r'\D', '', cnpj)
    if len(cnpj) != 14 or cnpj == cnpj[0]*14:
        return False
    pesos1 = [5,4,3,2,9,8,7,6,5,4,3,2]
    soma1 = sum(int(cnpj[i])*pesos1[i] for i in range(12))
    dig1 = 11 - (soma1 % 11)
    dig1 = 0 if dig1 >= 10 else dig1
    pesos2 = [6] + pesos1
    soma2 = sum(int(cnpj[i])*pesos2[i] for i in range(13))
    dig2 = 11 - (soma2 % 11)
    dig2 = 0 if dig2 >= 10 else dig2
    return dig1 == int(cnpj[12]) and dig2 == int(cnpj[13])

def salvar_excel(data):
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(PLANILHA_CAMPOS)
    else:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

    tipo = data.get('tipo_pessoa')
    cpf_cnpj = data.get('cpf', '') if tipo == 'pf' else data.get('cnpj', '')
    notes = f"CPF: {cpf_cnpj}" if tipo == 'pf' else f"CNPJ: {cpf_cnpj}"
    tipo_pessoa = "Pessoa Física" if tipo == 'pf' else "Pessoa Jurídica"
    nome = formatar_nome(data.get('nome', ''))
    nome_emoji = f"🔐 {nome}"
    celular = data.get('celular', '')
    celular_formatado = formatar_celular(celular)
    celular_link = celular_para_wa_link(celular)
    if celular_link:
        celular_excel = f'=HYPERLINK("{celular_link}"; "{celular_formatado}")'
    else:
        celular_excel = celular_formatado

    endereco = data.get('endereco', '').strip().title()
    numero = data.get('numero', '').strip()
    complemento = formatar_complemento(data.get('complemento', ''))
    bairro = data.get('bairro', '').strip().title()
    cidade = data.get('cidade', '').strip().title()
    cep = data.get('cep', '').strip()
    observacoes = formatar_observacao(data.get('observacao', ''))
    contato = data.get('contato', '').strip().title() if tipo == 'pj' else ""
    datahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    endereco_completo = f"{endereco}, {numero}"
    if complemento:
        endereco_completo += f" - {complemento}"
    endereco_completo += f" - {bairro}"

    ws.append([
        nome_emoji,             # Nome
        notes,                  # Notes (CPF/CNPJ)
        data.get('email', ''),  # E-mail
        celular_excel,          # Celular (link clicável com número formatado)
        endereco_completo,      # Endereço
        cidade,                 # Cidade
        cep,                    # CEP
        observacoes,            # Observações
        contato,                # Contato
        tipo_pessoa,            # Tipo de Pessoa
        datahora                # Data/Hora Cadastro
    ])
    wb.save(EXCEL_PATH)

@app.route('/', methods=['GET', 'POST'])
def formulario():
    if request.method == 'POST':
        data = request.form.to_dict()
        obrigatorios = ['tipo_pessoa', 'nome', 'email', 'celular', 'cep', 'endereco', 'bairro', 'cidade', 'numero']
        if data.get('tipo_pessoa') == 'pj':
            obrigatorios.append('contato')
        for campo in obrigatorios:
            if not data.get(campo):
                flash('Por favor, preencha todos os campos obrigatórios.', 'erro')
                return render_template('formulario.html', data=data)
        tipo = data.get('tipo_pessoa', '')
        # Validação de CPF/CNPJ
        if tipo == 'pf':
            cpf = data.get('cpf', '').strip()
            if not validar_cpf(cpf):
                data['erro_cpf'] = True
                flash('CPF inválido.', 'erro')
                return render_template('formulario.html', data=data)
        elif tipo == 'pj':
            cnpj = data.get('cnpj', '').strip()
            if not validar_cnpj(cnpj):
                data['erro_cnpj'] = True
                flash('CNPJ inválido.', 'erro')
                return render_template('formulario.html', data=data)
        # Formatação dos campos
        data['nome'] = formatar_nome(data.get('nome', ''))
        data['complemento'] = formatar_complemento(data.get('complemento', ''))
        data['observacao'] = formatar_observacao(data.get('observacao', ''))
        data['celular'] = formatar_celular(data.get('celular', ''))
        data['endereco'] = data.get('endereco', '').strip().title()
        data['bairro'] = data.get('bairro', '').strip().title()
        data['cidade'] = data.get('cidade', '').strip().title()
        salvar_excel(data)
        flash('Formulário enviado com sucesso! Em breve a Resolve Tecnologia entrará em contato com você.', 'sucesso')
        return redirect(url_for('formulario'))
    return render_template('formulario.html', data={})

@app.route('/registros')
def mostrar_registros():
    if not os.path.exists(EXCEL_PATH):
        return render_template('registros.html', registros=[], headers=PLANILHA_CAMPOS)
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0] if rows else []
    registros = rows[1:] if len(rows) > 1 else []
    return render_template('registros.html', headers=headers, registros=registros)

@app.template_filter('celular_link')
def celular_link(cell_value):
    if not cell_value:
        return ""
    s = str(cell_value)
    # Se for fórmula do Excel, extrai link e texto e monta o <a>
    if s.startswith('=HYPERLINK'):
        import re
        match = re.search(r'=HYPERLINK\("([^"]+)"\s*;\s*"([^"]+)"\)', s)
        if match:
            link = match.group(1)
            texto = match.group(2)
            return f'<a href="{link}" target="_blank">{texto}</a>'
        else:
            return s
    # Se for só o número, ainda faz o link
    digitos = ''.join(filter(str.isdigit, s))
    if len(digitos) >= 10:
        link = f"https://wa.me/55{digitos}"
        return f'<a href="{link}" target="_blank">{s}</a>'
    return s

if __name__ == '__main__':
    app.run(debug=True)